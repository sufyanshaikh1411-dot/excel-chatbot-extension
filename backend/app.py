from pathlib import Path
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import re

app = Flask(__name__)
CORS(app)

EXCEL_PATH = Path(__file__).resolve().parent.parent / "methodology.xlsx"

knowledge = []


def normalize_spaces(text):
    return re.sub(r"\s+", " ", str(text).strip())


def normalize_month_and_year(text):
    """
    Returns (month_name, year_4_digit) from strings like:
    Jan 26, Jan-26, January 2026, July,2025, Oct-2024
    """
    if not text:
        return None, None

    text = str(text).lower().strip()

    month_map = {
        "jan": "january",
        "january": "january",
        "feb": "february",
        "february": "february",
        "mar": "march",
        "march": "march",
        "apr": "april",
        "april": "april",
        "may": "may",
        "jun": "june",
        "june": "june",
        "jul": "july",
        "july": "july",
        "aug": "august",
        "august": "august",
        "sep": "september",
        "sept": "september",
        "september": "september",
        "oct": "october",
        "october": "october",
        "nov": "november",
        "november": "november",
        "dec": "december",
        "december": "december",
    }

    detected_month = None
    for key, value in month_map.items():
        if re.search(rf"\b{re.escape(key)}\b", text):
            detected_month = value
            break

    detected_year = None

    # 4-digit year first
    m4 = re.search(r"\b(20\d{2})\b", text)
    if m4:
        detected_year = m4.group(1)
    else:
        # 2-digit year like 26 => 2026
        m2 = re.search(r"\b(\d{2})\b", text)
        if m2:
            yy = int(m2.group(1))
            detected_year = f"20{yy:02d}"

    return detected_month, detected_year


def looks_like_month_group(text):
    """
    Detect if a cell is a month/year group header in column A.
    """
    if not text:
        return False

    month, year = normalize_month_and_year(text)
    return bool(month and year)


def load_excel():
    """
    Reads workbook and links Column A month/year group + month hyperlink
    to subsequent B/C rows until next month group appears.
    """
    global knowledge
    knowledge = []

    wb = load_workbook(EXCEL_PATH, data_only=True)

    for sheet in wb.worksheets:
        current_group_text = None
        current_group_link = None
        current_group_month = None
        current_group_year = None

        for row in sheet.iter_rows():
            col_a_text = ""
            col_a_link = None

            if len(row) >= 1:
                a_cell = row[0]
                if a_cell.value is not None:
                    col_a_text = normalize_spaces(a_cell.value)
                if a_cell.hyperlink and a_cell.hyperlink.target:
                    col_a_link = a_cell.hyperlink.target

            # If column A contains a month/year header, update current group
            if looks_like_month_group(col_a_text):
                current_group_text = col_a_text
                current_group_link = col_a_link
                current_group_month, current_group_year = normalize_month_and_year(col_a_text)

            row_values = []
            row_links = []

            for cell in row:
                if cell.value is not None:
                    text = normalize_spaces(cell.value)
                    if text:
                        row_values.append(text)

                if cell.hyperlink and cell.hyperlink.target:
                    row_links.append(cell.hyperlink.target)

            if not row_values and not row_links:
                continue

            joined_text = " | ".join(row_values)

            knowledge.append({
                "sheet": sheet.title,
                "text": joined_text,
                "links": row_links,
                "month_group": current_group_text,
                "month_link": current_group_link,
                "month": current_group_month,
                "year": current_group_year,
            })


load_excel()


def unique_items(items):
    unique = []
    seen = set()

    for item in items:
        key = (
            item["sheet"]
            + "||"
            + (item.get("month_group") or "")
            + "||"
            + item["text"]
            + "||"
            + "||".join(item.get("links", []))
        )
        if key not in seen:
            seen.add(key)
            unique.append(item)

    return unique


def format_results(items, limit=None):
    if not items:
        return "No relevant data found."

    output = []
    count = 0

    for item in items:
        if limit is not None and count >= limit:
            break

        block = []

        # Month header first
        if item.get("month_group"):
            if item.get("month_link"):
                block.append(f"{item['month_group']}\n{item['month_link']}")
            else:
                block.append(item["month_group"])

        # Main matched row
        block.append(item["text"])

        # Extra row-level links except duplicate month link
        extra_links = []
        for link in item.get("links", []):
            if link != item.get("month_link"):
                extra_links.append(link)

        if extra_links:
            block.extend(extra_links)

        output.append("\n".join(block))
        count += 1

    return "\n\n".join(output)


def score_keyword_match(text, keywords):
    score = 0
    text_lower = text.lower()

    for word in keywords:
        if word in text_lower:
            score += 1

    return score


def search_answer(question, selected_sheet):
    q = question.lower().strip()

    query_month, query_year = normalize_month_and_year(q)

    # Remove detected month/year tokens from keyword search
    cleaned_query = q

    month_words = [
        "jan", "january", "feb", "february", "mar", "march", "apr", "april",
        "may", "jun", "june", "jul", "july", "aug", "august", "sep", "sept",
        "september", "oct", "october", "nov", "november", "dec", "december"
    ]

    for word in month_words:
        cleaned_query = re.sub(rf"\b{re.escape(word)}\b", " ", cleaned_query)

    cleaned_query = re.sub(r"\b20\d{2}\b", " ", cleaned_query)
    cleaned_query = re.sub(r"\b\d{2}\b", " ", cleaned_query)

    keywords = [w.strip() for w in cleaned_query.split() if w.strip()]

    filtered = [
        item for item in knowledge
        if item["sheet"].strip().lower() == selected_sheet.strip().lower()
    ]

    if not filtered:
        return f"No data found for sheet '{selected_sheet}'."

    # CASE 1: exact month + year search => return ALL updates for that month group
    if query_month and query_year:
        month_matches = []

        for item in filtered:
            item_month = item.get("month")
            item_year = item.get("year")

            if item_month == query_month and item_year == query_year:
                month_matches.append(item)
                continue

            # fallback if month/year somehow also appears in text
            text_lower = item["text"].lower()
            short_year = query_year[-2:]
            if query_month in text_lower and (query_year in text_lower or re.search(rf"\b{re.escape(short_year)}\b", text_lower)):
                month_matches.append(item)

        month_matches = unique_items(month_matches)

        if month_matches:
            return format_results(month_matches, limit=None)

        return f"No updates found for {query_month.title()} {query_year} in sheet '{selected_sheet}'."

    # CASE 2: keyword search => return top matches, at least 5 if available
    scored = []

    for item in filtered:
        score = score_keyword_match(item["text"], keywords)

        # if no keywords left, fallback to generic whole query matching
        if not keywords:
            score = score_keyword_match(item["text"], q.split())

        if score > 0:
            scored.append((score, item))

    if scored:
        scored.sort(key=lambda x: x[0], reverse=True)
        ranked_items = [item for _, item in scored]
        ranked_items = unique_items(ranked_items)
        return format_results(ranked_items, limit=5)

    # CASE 3: generic fallback => top 5 based on all words
    fallback = []
    query_words = [w.strip() for w in q.split() if w.strip()]

    for item in filtered:
        score = score_keyword_match(item["text"], query_words)
        if score > 0:
            fallback.append((score, item))

    if fallback:
        fallback.sort(key=lambda x: x[0], reverse=True)
        fallback_items = [item for _, item in fallback]
        fallback_items = unique_items(fallback_items)
        return format_results(fallback_items, limit=5)

    return f"No relevant data found in sheet '{selected_sheet}'."


@app.route("/chat", methods=["POST"])
def chat():
    data = request.get_json(silent=True) or {}

    question = str(data.get("question", "")).strip()
    selected_sheet = str(data.get("sheet", "")).strip()

    if not question:
        return jsonify({"answer": "Please enter a question."})

    if not selected_sheet:
        return jsonify({"answer": "Please select a sheet."})

    answer = search_answer(question, selected_sheet)
    return jsonify({"answer": answer})


@app.route("/", methods=["GET"])
def home():
    return "Excel chatbot backend is running."


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
